from tkinter import *
import requests
from bs4 import BeautifulSoup
import openpyxl


window = Tk()

window.title("Mi APP")

#VARIABLES CONSTANTES 

url_pagina = 'https://www.statebcn.com/collections/tablas?page='

tablas_skate = {
    "url_imagen":"",
    "nombre": "",
    "marca": "",
    "precio_original": "",
    "precio_rebajado": "",
    "ahorro_porciento": "",
    "precio_original_rebajado": "",
}

def cargar_pagina(url):
    html = requests.get(url).content
    soup = BeautifulSoup(html, 'html.parser')
    return soup

def cargar_tablas(soup):
    tablas_list = list()
    list_elementos = soup.findAll("div",{"data-aos": "row-of-5"})

    i = 0

    for elemento in list_elementos:
        if i > 1:
            url_imagen = elemento.find("img", {"class": "grid-product__image"}).attrs["data-src"][2:-13]
            nombre = elemento.find("div", {"class": "grid-product__title grid-product__title--body"}).text
            marca = elemento.find("div", {"class": "grid-product__vendor"}).text
            if elemento.find("div", {"span": "grid-product__price--original"}) == None:
                precio_original = elemento.find("div", {"class": "grid-product__price"}).text.replace(" €","").replace("\n","")
                ahorro_porciento = "Sin ahorro"
                precio_rebajado = "Sin rebaja"
                precio_original_rebajado = "No tiene rebaja"
            else:
                precio_original = elemento.find("div", {"span": "grid-product__price--original"}).text.replace(" €", "")
                ahorro_porciento = elemento.find("span", {"class": "grid-product__price--savings"}).text.replace("\n", "").replace("%","").replace(" ", "")
                precio_rebajado = elemento.find("div", {"class": "grid-product__price"}).text
                precio_original_rebajado = elemento.find("span", {"class": "grid-product__price--original"}).text[:-2]

            # Diccionario
            skate = tablas_skate.copy()
            skate["url_imagen"] = url_imagen
            skate["nombre"] = nombre
            skate["marca"] = marca
            skate["precio_original"] = precio_original
            skate["precio_rebajado"] = precio_rebajado
            skate["ahorro_porciento"] = ahorro_porciento
            skate["precio_original_rebajado"] = precio_original_rebajado
            tablas_list.append(skate)
        i += 1

    return tablas_list


def cargar_todas_paginas():
    soup_list = list()
    for num_pag in range(1, 7):
        soup = cargar_pagina(url_pagina + str(num_pag))
        soup_list.append(soup)
    return soup_list

def cargar_todos_elementos():
    list_elementos = list()
    list_soup = cargar_todas_paginas()
    for soup in list_soup:
        list_tablas_paginas = cargar_tablas(soup)
        list_elementos.extend(list_tablas_paginas)
    return list_elementos

def exel_tablas():
    list_tablas = cargar_todos_elementos()
    documento_exel = openpyxl.Workbook()

    hoja1 = documento_exel.active

    hoja1.tittle = "tablas"

    count = 1

    for key in tablas_skate.keys():
        hoja1.cell(row=1, column=count, value=key)
        count += 1

    fila = 2

    for tabla in list_tablas:
        for columna in range(len(tablas_skate)):
            hoja1.cell(row=fila, column=columna + 1, value=tabla[list(tabla.keys())[columna]])
        fila += 1

    documento_exel.save("tablas.xlsx")

exel_tablas()